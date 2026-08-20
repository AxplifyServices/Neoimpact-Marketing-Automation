from __future__ import annotations

import os
from typing import Any, Iterator, List, Sequence, Tuple

import pandas as pd
from psycopg import sql

from app.storage.postgres_db import connection

DEFAULT_BATCH_SIZE = max(500, min(int(os.getenv("FILE_IMPORT_BATCH_SIZE", "5000")), 50000))
REQUIRED_COLUMNS = ("ID_Client", "Numero_Tel", "Mail")


def _normalize_columns(columns: Sequence[Any]) -> List[str]:
    return [str(column).strip() for column in columns]


def _iter_xlsx(path: str, batch_size: int) -> Iterator[pd.DataFrame]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        try:
            header = _normalize_columns(next(iterator))
        except StopIteration:
            yield pd.DataFrame()
            return
        rows: List[Tuple[Any, ...]] = []
        for row in iterator:
            rows.append(tuple(row))
            if len(rows) >= batch_size:
                yield pd.DataFrame.from_records(rows, columns=header)
                rows.clear()
        if rows:
            yield pd.DataFrame.from_records(rows, columns=header)
    finally:
        workbook.close()


def iter_file_batches(path: str, batch_size: int = DEFAULT_BATCH_SIZE) -> Iterator[pd.DataFrame]:
    size = max(500, min(int(batch_size or DEFAULT_BATCH_SIZE), 50000))
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        yield from pd.read_csv(path, chunksize=size)
        return
    if ext == ".xlsx":
        yield from _iter_xlsx(path, size)
        return
    if ext == ".xls":
        # Ancien format binaire : pandas/xlrd ne sait pas le streamer. Ce
        # fallback reste compatible mais doit être réservé aux petits fichiers.
        yield pd.read_excel(path, sheet_name=0)
        return
    if ext == ".parquet":
        import pyarrow.parquet as pq

        parquet = pq.ParquetFile(path)
        for batch in parquet.iter_batches(batch_size=size):
            yield batch.to_pandas()
        return
    if ext in {".jsonl", ".ndjson"}:
        yield from pd.read_json(path, lines=True, chunksize=size)
        return
    if ext == ".json":
        # JSON array classique non streamable par pandas ; JSONL est fortement
        # recommandé pour les gros volumes.
        yield pd.read_json(path)
        return
    raise ValueError("Type de fichier non supporté (csv/xlsx/xls/parquet/json/jsonl/ndjson)")


def _validate_batch(batch: pd.DataFrame, expected_columns: Sequence[str], *, first_batch: bool) -> pd.DataFrame:
    df = batch.copy(deep=False)
    df.columns = _normalize_columns(df.columns)
    expected = list(expected_columns)
    if first_batch:
        missing = sorted(set(expected) - set(df.columns))
        extra = sorted(set(df.columns) - set(expected))
        if missing or extra:
            messages = ["Fichier invalide (STRICT). Le schéma doit correspondre EXACTEMENT à la table clients."]
            if missing:
                messages.append(f"Colonnes manquantes ({len(missing)}) : {', '.join(missing)}")
            if extra:
                messages.append(f"Colonnes en trop ({len(extra)}) : {', '.join(extra)}")
            raise ValueError("\n".join(messages))
    for column in REQUIRED_COLUMNS:
        if column not in df.columns:
            raise ValueError(f"Fichier invalide : colonne obligatoire manquante : {column}")
        series = df[column]
        invalid = series.isna() | series.astype(str).str.strip().eq("")
        if bool(invalid.any()):
            raise ValueError(f"Fichier invalide : la colonne '{column}' contient des valeurs vides (obligatoire).")
    return df[expected]


def _to_db(value: Any) -> Any:
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return value


def bulk_import_clients(file_path: str, batch_size: int = DEFAULT_BATCH_SIZE) -> Tuple[int, int]:
    """Import strict, atomique et borné en mémoire.

    Le fichier est streamé vers une table TEMP PostgreSQL avec COPY. Les
    UPDATE/INSERT sont ensuite réalisés en SQL bulk. Aucune population complète
    n'est matérialisée dans Python.
    """
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT pg_advisory_xact_lock(%s)", (610002,))
            cur.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = 'public' AND table_name = 'clients'
                ORDER BY ordinal_position
                """
            )
            columns = [str(row[0]) for row in cur.fetchall()]
            if not columns:
                raise RuntimeError("Table clients introuvable.")
            lookup = {column.lower(): column for column in columns}
            key_column = lookup.get("id_client")
            radical_column = lookup.get("radical_compte")
            if not key_column or not radical_column:
                raise RuntimeError("Schéma clients invalide : ID_Client/radical_compte absents.")

            cur.execute("DROP TABLE IF EXISTS tmp_clients_import")
            cur.execute("CREATE TEMP TABLE tmp_clients_import ON COMMIT DROP AS SELECT * FROM clients WITH NO DATA")
            cur.execute('ALTER TABLE tmp_clients_import ADD COLUMN "__import_order" BIGSERIAL')

            copy_query = sql.SQL("COPY tmp_clients_import ({columns}) FROM STDIN").format(
                columns=sql.SQL(", ").join(sql.Identifier(column) for column in columns)
            )
            total_rows = 0
            first = True
            with cur.copy(copy_query) as copy:
                for raw_batch in iter_file_batches(file_path, batch_size=batch_size):
                    batch = _validate_batch(raw_batch, columns, first_batch=first)
                    first = False
                    for row in batch.itertuples(index=False, name=None):
                        copy.write_row(tuple(_to_db(value) for value in row))
                    total_rows += int(len(batch))

            if first:
                raise ValueError("Fichier vide.")

            # Dernière occurrence gagnante en cas d'ID_Client dupliqué dans le fichier.
            dedup = sql.SQL(
                """
                SELECT DISTINCT ON ({key}) {columns}
                FROM tmp_clients_import
                ORDER BY {key}, "__import_order" DESC
                """
            ).format(
                key=sql.Identifier(key_column),
                columns=sql.SQL(", ").join(sql.Identifier(column) for column in columns),
            )

            cur.execute(
                sql.SQL(
                    """
                    SELECT COUNT(*)
                    FROM ({dedup}) AS s
                    JOIN clients AS c ON c.{key} = s.{key}
                    """
                ).format(dedup=dedup, key=sql.Identifier(key_column))
            )
            row = cur.fetchone()
            existing_count = int(row[0] or 0) if row else 0

            update_columns = [column for column in columns if column not in {key_column, radical_column}]
            if update_columns:
                assignments = sql.SQL(", ").join(
                    sql.SQL("{col} = s.{col}").format(col=sql.Identifier(column))
                    for column in update_columns
                )
                distinct_left = sql.SQL(", ").join(sql.SQL("c.{col}").format(col=sql.Identifier(column)) for column in update_columns)
                distinct_right = sql.SQL(", ").join(sql.SQL("s.{col}").format(col=sql.Identifier(column)) for column in update_columns)
                cur.execute(
                    sql.SQL(
                        """
                        WITH s AS ({dedup})
                        UPDATE clients AS c
                        SET {assignments}
                        FROM s
                        WHERE c.{key} = s.{key}
                          AND ROW({left_values}) IS DISTINCT FROM ROW({right_values})
                        """
                    ).format(
                        dedup=dedup,
                        assignments=assignments,
                        key=sql.Identifier(key_column),
                        left_values=distinct_left,
                        right_values=distinct_right,
                    )
                )

            select_values = []
            for column in columns:
                if column == radical_column:
                    select_values.append(
                        sql.SQL(
                            "CASE WHEN NULLIF(BTRIM(COALESCE(s.{radical}::text, '')), '') IS NULL "
                            "THEN 'RC' || LPAD(nextval('client_radical_seq')::text, 8, '0') "
                            "ELSE s.{radical}::text END"
                        ).format(radical=sql.Identifier(radical_column))
                    )
                else:
                    select_values.append(sql.SQL("s.{col}").format(col=sql.Identifier(column)))

            cur.execute(
                sql.SQL(
                    """
                    WITH s AS ({dedup})
                    INSERT INTO clients ({columns})
                    SELECT {values}
                    FROM s
                    WHERE NOT EXISTS (
                        SELECT 1 FROM clients AS c WHERE c.{key} = s.{key}
                    )
                    ON CONFLICT ({key}) DO NOTHING
                    """
                ).format(
                    dedup=dedup,
                    columns=sql.SQL(", ").join(sql.Identifier(column) for column in columns),
                    values=sql.SQL(", ").join(select_values),
                    key=sql.Identifier(key_column),
                )
            )
            inserted = int(cur.rowcount if cur.rowcount is not None and cur.rowcount >= 0 else 0)

    # Compatibilité UI : "updated" signifie historiquement ID déjà présent,
    # même si les données étaient identiques et que l'UPDATE physique a été évité.
    return inserted, existing_count


def materialize_file_members(id_cible: str, file_path: str, batch_size: int = DEFAULT_BATCH_SIZE, *, replace: bool = False) -> int:
    """Matérialise seulement les clés de la cible fichier dans clients_cibles.

    Le fichier peut ensuite être conservé pour audit, mais les campagnes ne le
    relisent plus et ne construisent plus de DataFrame complet.
    """
    cible_id = str(id_cible or "").strip()
    if not cible_id:
        raise ValueError("id_cible manquant")

    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("CREATE TEMP TABLE tmp_file_target_ids (client_id TEXT NOT NULL) ON COMMIT DROP")
            copy_query = "COPY tmp_file_target_ids (client_id) FROM STDIN"
            first = True
            with cur.copy(copy_query) as copy:
                for raw_batch in iter_file_batches(file_path, batch_size=batch_size):
                    raw_batch.columns = _normalize_columns(raw_batch.columns)
                    if first and "ID_Client" not in raw_batch.columns:
                        raise ValueError("Fichier cible invalide : colonne ID_Client manquante.")
                    first = False
                    for value in raw_batch["ID_Client"].tolist():
                        if value is None or (isinstance(value, float) and pd.isna(value)):
                            continue
                        client_id = str(value).strip()
                        if client_id:
                            copy.write_row((client_id,))
            if first:
                raise ValueError("Fichier vide.")

            if replace:
                cur.execute('DELETE FROM clients_cibles WHERE "ID_CIBLE" = %s', (cible_id,))

            cur.execute(
                """
                INSERT INTO clients_cibles ("ID_CIBLE", "Radical_compte", created_at)
                SELECT %s, c.radical_compte, NOW()::text
                FROM (SELECT DISTINCT client_id FROM tmp_file_target_ids) AS f
                JOIN clients AS c ON c."ID_Client"::text = f.client_id
                WHERE BTRIM(COALESCE(c.radical_compte::text, '')) <> ''
                ON CONFLICT ("ID_CIBLE", "Radical_compte") DO NOTHING
                """,
                (cible_id,),
            )
            return int(cur.rowcount if cur.rowcount is not None and cur.rowcount >= 0 else 0)
